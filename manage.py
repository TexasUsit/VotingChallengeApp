from VotingApp.mainApp import app, db, update_ret, add_stock, update_score, create_stock_info
from VotingApp.models import User, Tickers, Role, Transactions
from flask_script import Manager, prompt_bool
from openpyxl import load_workbook, Workbook
import requests

manager = Manager(app)

@manager.command
def initdb():
    db.create_all()
    admin = Role(name="Admin", description="Gets to look at all the rankings.")
    db.session.add(User(firstName="Arnav", lastName="Jain",email="arnav_jain@utexas.edu", password="test11", roles=[admin]))
    db.session.commit()
    refreshdb()
    print 'Initialized the database'

@manager.command
def dropdb():
    if prompt_bool(
        "Are you sure you want to lose all your data"):
        db.drop_all()
        print 'Dropped the database'

@manager.command
def refreshdb():
    # Refresh the score and ranks for each student
    for student in User.query.all():
        numStocks = update_ret(student, student.stocks, student.transactions)
        update_score(student, student.ret, numStocks)

    # Refresh the stored information for each stock
    for stock in Tickers.query.all():
        create_stock_info(stock)
    for stock in Transactions.query.all():
        create_stock_info(stock)


@manager.command
def addstock():
    name = str(raw_input("What is the file name? "))
    symbol = str(raw_input("Ticker: "))
    price = int(raw_input("Starting price: $"))
    num_votes = int(raw_input("How many votes were there? "))

    wb = load_workbook(filename=name)
    ws = wb.active

    for index in range(1, num_votes + 1):
        if str(ws['B' + str(index)].value) == 'Long':
            if  Tickers.query.filter_by(short = False, ticker = symbol).count() > 0:
                stock = Tickers.query.filter_by(short = False, ticker = symbol).first()
                print("I FOUND THE STOCK ALREADY")
            else:
                stock = Tickers(ticker=symbol, startingPrice=price, short=False)
        elif str(ws['B' + str(index)].value) == 'Short':
            if Tickers.query.filter_by(short = True, ticker = symbol).count() > 0:
                stock = Tickers.query.filter_by(short = True, ticker = symbol).first()
                print("I FOUND THE STOCK ALREADY")
            else:
                stock = Tickers(ticker=symbol, startingPrice=price, short=True)

        if User.query.filter_by(email=str(ws['A'+str(index)].value)) != None:
            student = User.query.filter_by(email=str(ws['A'+str(index)].value)).first()
            add_stock(student, stock)

@manager.command
def fixdb():
    num_students = User.query.count()
    num_students_wo_id = 1
    for student in User.query.all():
        if student.id == None:
            student.id == num_students + num_students_wo_id
            num_students_wo_id += 1
            print("I fixed", student.firstName, "!")

        db.session.commit()

if __name__ == '__main__':
    manager.run()


